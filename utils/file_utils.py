# utils/file_utils.py
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from datetime import datetime
import sqlite3
from utils.config_loader import CONFIG
from utils.date_utils import format_date_for_display

def export_agents_to_excel(main_window, db_manager):
    """Exporte la liste complète des agents vers un fichier Excel."""
    main_window.config(cursor="watch")
    main_window.update_idletasks()
    main_window.set_status("Exportation des agents en cours...")
    
    try:
        # --- MODIFICATION APPLIQUÉE ICI ---
        # On récupère tous les agents sans pagination pour un export complet
        agents = db_manager.get_agents()
        if not agents:
            messagebox.showinfo("Information", "Aucun agent à exporter.")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Fichiers Excel", "*.xlsx")],
            title="Exporter la liste des agents",
            initialfile=f"Export_Agents_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        )
        if not filename: return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Agents"
        headers = ["ID", "Nom", "Prénom", "PPR", "Grade", "Solde"]
        ws.append(headers)
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for agent in agents:
            # On transforme l'objet Agent en une liste de valeurs pour l'écriture
            ws.append([agent.id, agent.nom, agent.prenom, agent.ppr, agent.grade, agent.solde])

        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value or "")) for cell in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
        
        wb.save(filename)
        messagebox.showinfo("Succès", f"Liste des agents exportée avec succès vers\n{filename}")
    except Exception as e:
        messagebox.showerror("Erreur d'écriture", f"Impossible de sauvegarder le fichier : {e}")
    finally:
        main_window.config(cursor="")
        main_window.set_status("Prêt.")

def export_all_conges_to_excel(main_window, db_manager):
    """Exporte la liste complète de tous les congés vers un fichier Excel."""
    main_window.config(cursor="watch")
    main_window.update_idletasks()
    main_window.set_status("Exportation totale en cours...")
    
    try:
        conges = db_manager.get_conges()
        agents_cache = {agent.id: agent for agent in db_manager.get_agents()} # Cache pour éviter les requêtes répétées

        if not conges:
            messagebox.showinfo("Information", "Aucun congé à exporter.")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Fichiers Excel", "*.xlsx")],
            title="Exporter tous les congés",
            initialfile=f"Export_Conges_Total_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        )
        if not filename: return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tous les Congés"
        headers = ["PPR Agent", "Nom Agent", "Prénom Agent", "Type Congé", "Début", "Fin", "Jours Pris", "Statut", "Justification", "Intérimaire"]
        ws.append(headers)
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for c in conges:
            agent = agents_cache.get(c.agent_id)
            if not agent: continue # Agent supprimé, on ignore son congé dans l'export

            interim_info = ""
            if c.interim_id:
                interim = agents_cache.get(c.interim_id)
                interim_info = f"{interim.nom} {interim.prenom}" if interim else "Agent Supprimé"
            
            row_data = [
                agent.ppr, agent.nom, agent.prenom, c.type_conge, 
                format_date_for_display(c.date_debut), 
                format_date_for_display(c.date_fin), 
                c.jours_pris, c.statut, c.justif or "", interim_info
            ]
            ws.append(row_data)

        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value or "")) for cell in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
            
        wb.save(filename)
        messagebox.showinfo("Succès", f"Tous les congés ont été exportés avec succès vers\n{filename}")
    except Exception as e:
        messagebox.showerror("Erreur d'écriture", f"Impossible de sauvegarder le fichier : {e}")
    finally:
        main_window.config(cursor="")
        main_window.set_status("Prêt.")

def import_agents_from_excel(main_window, db_manager):
    """Importe des agents depuis un fichier Excel. Seuls le nom et le prénom sont obligatoires."""
    filename = filedialog.askopenfilename(
        title="Sélectionner un fichier Excel à importer",
        filetypes=[("Fichiers Excel", "*.xlsx")]
    )
    if not filename:
        return

    main_window.config(cursor="watch")
    main_window.update_idletasks()
    main_window.set_status("Importation en cours...")
    
    errors = []
    # --- MODIFICATION APPLIQUÉE ICI ---
    # On ne vérifie plus les en-têtes optionnels, seulement les obligatoires.
    agent_import_headers_obligatoires = CONFIG.get('agent_import_headers_obligatoires', ['nom', 'prenom'])
    grades = CONFIG['ui']['grades']

    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        header = [str(cell.value).lower().strip() for cell in ws[1] if cell.value]
        
        if not all(h in header for h in agent_import_headers_obligatoires):
            raise ValueError(f"Le fichier Excel doit contenir au minimum les colonnes : {', '.join(agent_import_headers_obligatoires)}")

        col_map = {name: i for i, name in enumerate(header)}
        added_count, updated_count, error_count = 0, 0, 0
        
        db_manager.conn.execute('BEGIN TRANSACTION')
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if all(c is None for c in row): continue
                
                # Récupération des champs obligatoires
                nom = str(row[col_map['nom']]).strip()
                prenom = str(row[col_map['prenom']]).strip()
                if not nom or not prenom:
                    raise ValueError("Les colonnes 'nom' et 'prenom' ne peuvent pas être vides.")

                # Récupération des champs optionnels
                ppr_val = row[col_map['ppr']] if 'ppr' in col_map else None
                ppr = str(ppr_val).strip() if ppr_val and str(ppr_val).strip() else None

                grade_val = row[col_map['grade']] if 'grade' in col_map else None
                grade = str(grade_val).strip() if grade_val and str(grade_val).strip() else None
                if grade and grade not in grades: 
                    raise ValueError(f"Grade '{grade}' invalide.")

                solde_val = row[col_map['solde']] if 'solde' in col_map else None
                solde = float(str(solde_val).replace(",", ".")) if solde_val is not None and str(solde_val).strip() != "" else 0.0
                if solde < 0: 
                    raise ValueError(f"Le solde '{solde}' ne peut être négatif.")

                agent = db_manager.get_agent_by_ppr(ppr)
                if agent:
                    if not db_manager.modifier_agent(agent[0], nom, prenom, ppr, grade, solde):
                        raise sqlite3.Error(f"Erreur de mise à jour pour PPR {ppr}.")
                    updated_count += 1
                else:
                    if not db_manager.ajouter_agent(nom, prenom, ppr, grade, solde):
                        # Si l'ajout échoue (par exemple, PPR déjà pris par un autre agent), on le signale.
                        raise sqlite3.Error(f"Erreur d'ajout pour {nom} {prenom} (PPR: {ppr}). Le PPR est peut-être déjà utilisé.")
                    added_count += 1
            except (ValueError, TypeError, IndexError) as ve:
                errors.append(f"Ligne {i}: {ve}"); error_count += 1
            except Exception as e:
                errors.append(f"Ligne {i}: Erreur BD - {e}"); error_count += 1
        
        if error_count > 0:
            raise Exception("Des erreurs ont été détectées. L'importation est annulée.")
        
        db_manager.conn.commit()
        summary = f"Importation réussie !\n\n- Agents ajoutés : {added_count}\n- Agents mis à jour : {updated_count}"
        messagebox.showinfo("Rapport d'importation", summary)
    except Exception as e:
        if db_manager.conn.in_transaction:
            db_manager.conn.rollback()
        summary = f"Échec de l'importation: {e}\n\nAucune modification n'a été enregistrée."
        if errors:
            summary += "\n\nDétail des erreurs (premières 5):\n" + "\n".join(errors[:5])
        messagebox.showerror("Rapport d'importation", summary)
    finally:
        main_window.config(cursor="")
        main_window.set_status("Prêt.")
        main_window.refresh_all()